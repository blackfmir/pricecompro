from __future__ import annotations

from io import BytesIO
from typing import Any

from lxml import etree
from openpyxl import load_workbook

from app.schemas.supplier_product import SupplierProductCreate


def _col_letter_to_index(letter: str) -> int:
    """A->1, B->2, Z->26, AA->27 ... (1-based)."""
    s = letter.strip().upper()
    n = 0
    for ch in s:
        if not ("A" <= ch <= "Z"):
            return 0
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _to_str(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s != "" else None


def _to_float(s: str | None) -> float | None:
    if s is None:
        return None
    s = s.replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def _split_list(s: str | None, sep: str | None) -> list[str] | None:
    if not s or not sep:
        return None
    parts = [x.strip() for x in s.split(sep)]
    return [x for x in parts if x]


def import_xlsx_bytes(
    *,
    file_bytes: bytes,
    supplier_id: int,
    price_list_id: int,
    mapping: dict[str, Any],
    source_config: dict[str, Any] | None,
) -> tuple[list[SupplierProductCreate], list[str]]:
    """
    Підтримка by: col_letter | col_index | header_name
    source_config:
      {
        "sheet": {"by":"name","value":"Прайс"} | {"by":"index","value":1},
        "header": false,           # якщо True — читаємо імена з рядка start_row
        "start_row": 1
      }
    """
    sc = source_config or {}
    sheet_spec = sc.get("sheet") or {"by": "index", "value": 1}
    header = bool(sc.get("header", False))
    start_row = int(sc.get("start_row", 1))

    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    if sheet_spec.get("by") == "name":
        ws = wb[sheet_spec.get("value")]
    else:
        idx = int(sheet_spec.get("value", 1)) - 1
        ws = wb.worksheets[idx]

    header_map: dict[str, int] = {}
    data_start = start_row
    if header:
        # будуємо мапу "header_name_lower" -> col_index (1-based)
        hdr_row = ws[start_row]
        for i, cell in enumerate(hdr_row, start=1):
            name = _to_str(cell.value)
            if name:
                header_map[name.lower()] = i
        data_start = start_row + 1

    def read_cell(row_idx: int, spec: dict[str, Any] | str | None) -> str | None:
        if not spec:
            return None
        # зворотна сумісність: просто рядок означає header_name
        if isinstance(spec, str):
            name = spec.strip().lower()
            col = header_map.get(name)
            if not col:
                return None
            return _to_str(ws.cell(row=row_idx, column=col).value)

        by = (spec.get("by") or "").strip().lower()
        if by == "col_letter":
            col = _col_letter_to_index(str(spec.get("value")))
            if col <= 0:
                return None
            return _to_str(ws.cell(row=row_idx, column=col).value)
        if by == "col_index":
            col = int(spec.get("value", 0))
            if col <= 0:
                return None
            return _to_str(ws.cell(row=row_idx, column=col).value)
        if by == "header_name":
            name = str(spec.get("value", "")).strip().lower()
            col = header_map.get(name)
            if not col:
                return None
            return _to_str(ws.cell(row=row_idx, column=col).value)
        return None

    items: list[SupplierProductCreate] = []
    errors: list[str] = []

    max_row = ws.max_row or 0
    for r in range(data_start, max_row + 1):
        try:
            supplier_sku = read_cell(r, mapping.get("supplier_sku")) or ""
            if not supplier_sku:
                # порожні рядки пропускаємо
                continue

            def opt_sep(key: str) -> str | None:
                spec = mapping.get(key) or {}
                if isinstance(spec, dict):
                    opts = spec.get("options") or {}
                    return opts.get("split")
                return None

            item = SupplierProductCreate(
                supplier_id=supplier_id,
                price_list_id=price_list_id,
                supplier_sku=supplier_sku,
                manufacturer_sku=read_cell(r, mapping.get("manufacturer_sku")),
                mpn=read_cell(r, mapping.get("mpn")),
                gtin=read_cell(r, mapping.get("gtin")),
                ean=read_cell(r, mapping.get("ean")),
                upc=read_cell(r, mapping.get("upc")),
                jan=read_cell(r, mapping.get("jan")),
                isbn=read_cell(r, mapping.get("isbn")),
                name=read_cell(r, mapping.get("name")),
                brand_raw=read_cell(r, mapping.get("brand_raw")),
                category_raw=read_cell(r, mapping.get("category_raw")),
                price_raw=_to_float(read_cell(r, mapping.get("price_raw"))),
                currency_raw=read_cell(r, mapping.get("currency_raw")),
                qty_raw=_to_float(read_cell(r, mapping.get("qty_raw"))),
                availability_text=read_cell(r, mapping.get("availability_text")),
                delivery_terms=read_cell(r, mapping.get("delivery_terms")),
                delivery_date=read_cell(r, mapping.get("delivery_date")),  # парсинг дат додамо окремо, якщо треба
                location=read_cell(r, mapping.get("location")),
                short_description_raw=read_cell(r, mapping.get("short_description_raw")),
                description_raw=read_cell(r, mapping.get("description_raw")),
                image_urls=_split_list(read_cell(r, mapping.get("image_urls")), opt_sep("image_urls")),
            )
            items.append(item)
        except Exception as e:
            errors.append(f"Row {r}: {e!r}")

    return items, errors


def _xpath_text(node, expr: str, ns: dict[str, str] | None) -> str | None:
    res = node.xpath(expr, namespaces=ns)
    if not res:
        return None
    val = res[0]
    if isinstance(val, etree._Element):
        return (val.text or "").strip() or None
    return (str(val).strip() or None)

def _xpath_list(node, expr: str, ns: dict[str, str] | None) -> list[str] | None:
    res = node.xpath(expr, namespaces=ns) or []
    out: list[str] = []
    for v in res:
        t = ((v.text or "") if isinstance(v, etree._Element) else str(v)).strip()
        if t:
            out.append(t)
    return out or None

def _read_field(node, spec, ns) -> str | None:
    if not spec:
        return None
    if isinstance(spec, str):
        # скорочення: трактуємо як xpath
        return _xpath_text(node, spec, ns)
    by = (spec.get("by") or "").lower()
    val = spec.get("value")
    if by in {"xpath", "text"}:
        return _xpath_text(node, str(val), ns)
    if by == "attr":
        if not isinstance(val, str) or not val.startswith("@"):
            return None
        attr = val[1:]
        raw = node.get(attr)
        return raw.strip() if isinstance(raw, str) else None
    if by == "xpath_list":
        # повернемо як JSON-закодований список вище по логіці (список обробимо окремо)
        return "\u0000LIST\u0000" + "|".join(_xpath_list(node, str(val), ns) or [])
    return None

def import_xml_bytes(
    *,
    file_bytes: bytes,
    supplier_id: int,
    price_list_id: int,
    mapping: dict,
    source_config: dict | None,
) -> tuple[list[SupplierProductCreate], list[str]]:
    """
    source_config:
      {
        "namespaces": {"y":"yml"},
        "containers": {
           "items": "/yml_catalog/shop/offers/offer",
           "categories": "/yml_catalog/shop/categories/category"
        }
      }
    mapping: {
      "product_fields": {...}, "category_fields": {...}   # новий формат
      # або плоский dict для сумісності: трактуємо як product_fields
    }
    """
    sc = source_config or {}
    ns = sc.get("namespaces") or {}
    cont = (sc.get("containers") or {})
    items_xpath = cont.get("items")
    if not items_xpath:
        return [], ["containers.items is required for XML"]

    root = etree.fromstring(file_bytes)

    # читаємо довідник категорій (опційно)
    cats: dict[str, dict] = {}
    cat_xpath = cont.get("categories")
    cat_fields = (mapping.get("category_fields") if isinstance(mapping, dict) else None) or {}
    if cat_xpath and cat_fields:
        for cat in root.xpath(cat_xpath, namespaces=ns) or []:
            cid = _read_field(cat, cat_fields.get("id"), ns)
            if not cid:
                continue
            cats[cid] = {
                "name": _read_field(cat, cat_fields.get("name"), ns),
                "parent_id": _read_field(cat, cat_fields.get("parent_id"), ns),
            }

    # готуємо перелік товарів
    product_fields = (mapping.get("product_fields") if isinstance(mapping, dict) else mapping) or {}
    items_nodes = root.xpath(items_xpath, namespaces=ns) or []

    out: list[SupplierProductCreate] = []
    errs: list[str] = []

    for n in items_nodes:
        try:
            supplier_sku = _read_field(n, product_fields.get("supplier_sku"), ns) or ""
            if not supplier_sku:
                continue

            def read(name: str, _n=n) -> str | None:
                v = _read_field(_n, product_fields.get(name), ns)
                if isinstance(v, str) and v.startswith("\u0000LIST\u0000"):
                    return v  # маркер списку
                return v

            image_field = read("image_urls")
            images: list[str] | None = None
            if image_field and image_field.startswith("\u0000LIST\u0000"):
                images = [x for x in image_field.replace("\u0000LIST\u0000", "", 1).split("|") if x]

            cat_id_ref = read("category_id_ref")
            category_raw = read("category_raw")
            if (not category_raw) and cat_id_ref and cat_id_ref in cats:
                category_raw = cats[cat_id_ref].get("name")

            def fnum(x):  # float parser
                if not x:
                    return None
                x = x.replace(" ", "").replace(",", ".")
                try:
                    return float(x)
                except Exception:
                    return None

            item = SupplierProductCreate(
                supplier_id=supplier_id,
                price_list_id=price_list_id,
                supplier_sku=supplier_sku,
                manufacturer_sku=read("manufacturer_sku"),
                mpn=read("mpn"),
                gtin=read("gtin"),
                ean=read("ean"),
                upc=read("upc"),
                jan=read("jan"),
                isbn=read("isbn"),
                name=read("name"),
                brand_raw=read("brand_raw"),
                category_raw=category_raw,
                price_raw=fnum(read("price_raw")),
                currency_raw=read("currency_raw"),
                qty_raw=fnum(read("qty_raw")),
                availability_text=read("availability_text"),
                delivery_terms=read("delivery_terms"),
                delivery_date=read("delivery_date"),
                location=read("location"),
                short_description_raw=read("short_description_raw"),
                description_raw=read("description_raw"),
                image_urls=images,
            )
            out.append(item)
        except Exception as e:
            errs.append(f"XML item error: {e!r}")

    return out, errs
